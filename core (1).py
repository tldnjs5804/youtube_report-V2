# -*- coding: utf-8 -*-
"""
core.py

YouTube Data API v3를 이용해 채널의 지정 기간 업로드 영상을 수집하고
엑셀로 정리하는 핵심 로직. (yt-dlp 크롤링 방식에서 공식 API 방식으로 전환)

전환 이유:
- yt-dlp 크롤링은 유튜브의 봇 탐지, DRM 스트림 제한, 연령 제한 등으로
  클라우드 서버(Streamlit Cloud) 환경에서 불안정했음.
- 공식 API는 안정적이고, 조회수/좋아요/댓글수 등 정확한 통계를 제공함.

구성:
1. 채널 식별: @핸들/커스텀URL/채널ID 등 다양한 입력을 channelId로 정규화
2. 업로드 목록: channels.list로 업로드 재생목록ID 획득 → playlistItems.list로
   전체 영상 ID 수집 (페이지네이션)
3. 상세 정보: videos.list로 50개씩 배치 조회 (조회수/좋아요/댓글/길이/설명/라이브여부)
4. 롱폼/숏폼 판별: YouTube Data API에는 "이 영상이 Shorts인지" 알려주는
   공식 필드가 없음. 재생시간 추측 대신, /shorts/{video_id} URL에 리다이렉트
   없이 HEAD 요청을 보내 실제 유튜브의 판정을 확인한다.
   - HTTP 200 -> Shorts로 분류된 영상 (그대로 재생됨)
   - HTTP 3xx(리다이렉트) -> 일반(롱폼) 영상 (watch 페이지로 리다이렉트됨)
5. 라이브 판별: videos.list의 liveStreamingDetails 필드 존재 여부로 확정
   (진행중/예정/종료된 라이브 방송에만 이 필드가 존재함 — 공식 스펙 기반이라 100% 정확)
6. 번역: 설명이 외국어로 판단되면 deep-translator로 한국어 번역 (선택)
7. 엑셀: 전체/롱폼/숏폼/라이브 4개 시트, 다중 채널은 채널별 시트 + 통합 시트
"""

import re
import time
from datetime import datetime, timezone

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

try:
    from deep_translator import GoogleTranslator
    TRANSLATOR_AVAILABLE = True
except ImportError:
    TRANSLATOR_AVAILABLE = False


API_BASE = "https://www.googleapis.com/youtube/v3"
SUMMARY_MAX_CHARS = 200


class YouTubeAPIError(Exception):
    """YouTube Data API 호출 중 발생한 오류. 사용자에게 보여줄 메시지를 담는다."""
    pass


# ---------------------------------------------------------------------------
# 채널 식별
# ---------------------------------------------------------------------------

def extract_handle_or_id(raw: str) -> dict:
    """
    사용자 입력(핸들, URL, 채널ID 등)에서 channels.list 호출에 쓸 파라미터를 뽑아낸다.
    반환값: {"forHandle": ...} 또는 {"id": ...} 또는 {"forUsername": ...} 중 하나.
    """
    raw = raw.strip()

    # 순수 채널 ID 형식 (UC로 시작, 24자)
    if re.fullmatch(r"UC[\w-]{22}", raw):
        return {"id": raw}

    # URL에서 부분 추출
    if raw.startswith("http://") or raw.startswith("https://"):
        m = re.search(r"youtube\.com/channel/(UC[\w-]{22})", raw)
        if m:
            return {"id": m.group(1)}

        m = re.search(r"youtube\.com/@([\w.\-]+)", raw)
        if m:
            return {"forHandle": "@" + m.group(1)}

        m = re.search(r"youtube\.com/c/([\w.\-]+)", raw)
        if m:
            return {"forUsername_or_handle": m.group(1)}  # 레거시, 아래에서 순차 시도

        m = re.search(r"youtube\.com/user/([\w.\-]+)", raw)
        if m:
            return {"forUsername": m.group(1)}

        # youtube.com/이름 (탭도 /c/도 없는 가장 오래된 형식)
        m = re.search(r"youtube\.com/([\w.\-]+)/?$", raw)
        if m and m.group(1) not in ("watch", "shorts", "channel", "playlist"):
            return {"forUsername_or_handle": m.group(1)}

        raise YouTubeAPIError(f"채널 주소에서 채널 정보를 찾을 수 없습니다: {raw}")

    if raw.startswith("@"):
        return {"forHandle": raw}

    # 핸들도 URL도 아니면 @를 붙여 핸들로 간주
    return {"forHandle": "@" + raw}


def resolve_channel_id(raw: str, api_key: str) -> dict:
    """
    사용자 입력을 실제 channelId와 업로드 재생목록ID로 변환한다.
    반환값: {"channel_id": ..., "uploads_playlist_id": ..., "title": ...}
    """
    params_variants = extract_handle_or_id(raw)

    # forUsername_or_handle은 순서대로 여러 방식을 시도해야 함
    if "forUsername_or_handle" in params_variants:
        name = params_variants["forUsername_or_handle"]
        attempts = [
            {"forHandle": "@" + name},
            {"forUsername": name},
        ]
    else:
        attempts = [params_variants]

    last_error = None
    for attempt_params in attempts:
        try:
            resp = requests.get(
                f"{API_BASE}/channels",
                params={
                    "part": "contentDetails,snippet",
                    "key": api_key,
                    **attempt_params,
                },
                timeout=15,
            )
            data = resp.json()

            if "error" in data:
                last_error = data["error"].get("message", str(data["error"]))
                continue

            items = data.get("items", [])
            if not items:
                last_error = "채널을 찾을 수 없음"
                continue

            channel = items[0]
            uploads_playlist_id = channel["contentDetails"]["relatedPlaylists"]["uploads"]
            return {
                "channel_id": channel["id"],
                "uploads_playlist_id": uploads_playlist_id,
                "title": channel["snippet"]["title"],
            }
        except requests.RequestException as e:
            last_error = str(e)
            continue

    # 마지막 수단: search.list로 채널명 검색 (쿼터를 더 쓰지만 가장 관대함)
    try:
        search_query = raw.strip().lstrip("@").split("/")[-1]
        resp = requests.get(
            f"{API_BASE}/search",
            params={
                "part": "snippet",
                "q": search_query,
                "type": "channel",
                "maxResults": 1,
                "key": api_key,
            },
            timeout=15,
        )
        data = resp.json()
        if "error" in data:
            raise YouTubeAPIError(data["error"].get("message", "채널 검색 실패"))

        items = data.get("items", [])
        if items:
            channel_id = items[0]["snippet"]["channelId"]
            # 채널ID로 다시 정확히 조회
            resp2 = requests.get(
                f"{API_BASE}/channels",
                params={"part": "contentDetails,snippet", "id": channel_id, "key": api_key},
                timeout=15,
            )
            data2 = resp2.json()
            items2 = data2.get("items", [])
            if items2:
                channel = items2[0]
                uploads_playlist_id = channel["contentDetails"]["relatedPlaylists"]["uploads"]
                return {
                    "channel_id": channel["id"],
                    "uploads_playlist_id": uploads_playlist_id,
                    "title": channel["snippet"]["title"],
                }
    except requests.RequestException as e:
        last_error = str(e)

    raise YouTubeAPIError(f"'{raw}' 채널을 찾을 수 없습니다 ({last_error}). 주소를 확인해주세요.")


# ---------------------------------------------------------------------------
# 업로드 목록 + 상세 정보 수집
# ---------------------------------------------------------------------------

def fetch_playlist_video_ids(uploads_playlist_id: str, api_key: str,
                              start_date: datetime, end_date: datetime,
                              progress_callback=None):
    """
    업로드 재생목록에서 [start_date, end_date] 범위의 영상 ID를 수집한다.
    playlistItems.list는 최신순으로 반환하므로, start_date보다 오래된 영상이
    나오면 조기 종료한다.
    """
    video_ids = []
    page_token = None

    while True:
        params = {
            "part": "contentDetails",
            "playlistId": uploads_playlist_id,
            "maxResults": 50,
            "key": api_key,
        }
        if page_token:
            params["pageToken"] = page_token

        resp = requests.get(f"{API_BASE}/playlistItems", params=params, timeout=15)
        data = resp.json()

        if "error" in data:
            raise YouTubeAPIError(data["error"].get("message", "재생목록 조회 실패"))

        items = data.get("items", [])
        stop = False

        for item in items:
            published_at_str = item["contentDetails"].get("videoPublishedAt")
            if not published_at_str:
                continue
            published_at = datetime.fromisoformat(published_at_str.replace("Z", "+00:00"))
            published_at_naive = published_at.replace(tzinfo=None)

            if published_at_naive > end_date:
                continue
            if published_at_naive < start_date:
                stop = True
                break

            video_ids.append(item["contentDetails"]["videoId"])

        if progress_callback:
            progress_callback(f"      -> 지금까지 {len(video_ids)}개 영상 ID 수집...")

        if stop:
            break

        page_token = data.get("nextPageToken")
        if not page_token:
            break

    return video_ids


def fetch_videos_detail(video_ids: list, api_key: str):
    """
    videos.list를 50개씩 배치로 호출해 상세 정보를 가져온다.
    """
    all_items = []
    for i in range(0, len(video_ids), 50):
        batch = video_ids[i:i + 50]
        resp = requests.get(
            f"{API_BASE}/videos",
            params={
                "part": "snippet,contentDetails,statistics,liveStreamingDetails",
                "id": ",".join(batch),
                "key": api_key,
            },
            timeout=15,
        )
        data = resp.json()

        if "error" in data:
            raise YouTubeAPIError(data["error"].get("message", "영상 상세 조회 실패"))

        all_items.extend(data.get("items", []))

    return all_items


# ---------------------------------------------------------------------------
# 숏폼 판별 (재생시간 추측이 아니라 실제 /shorts/ URL 응답 확인)
# ---------------------------------------------------------------------------

def is_shorts_video(video_id: str) -> bool:
    """
    https://www.youtube.com/shorts/{video_id} 로 리다이렉트 없이 요청을 보내
    실제 유튜브가 이 영상을 Shorts로 분류하는지 확인한다.
    - HTTP 200 -> Shorts
    - 그 외(리다이렉트 등) -> 롱폼
    네트워크 오류 시 판별 불가로 처리하고(롱폼 취급), 호출부에서 재생시간을
    보조 판단 근거로 쓸 수 있게 None을 반환한다.
    """
    url = f"https://www.youtube.com/shorts/{video_id}"
    try:
        resp = requests.head(url, allow_redirects=False, timeout=8,
                              headers={"User-Agent": "Mozilla/5.0"})
        if resp.status_code == 200:
            return True
        if 300 <= resp.status_code < 400:
            return False
        return None
    except requests.RequestException:
        return None


# ---------------------------------------------------------------------------
# 번역 / 요약
# ---------------------------------------------------------------------------

def is_mostly_korean(text: str, threshold: float = 0.3) -> bool:
    if not text:
        return True
    letters = [ch for ch in text if ch.isalpha()]
    if not letters:
        return True
    korean_count = sum(1 for ch in letters if "\uac00" <= ch <= "\ud7a3")
    return (korean_count / len(letters)) >= threshold


def translate_to_korean(text: str) -> str:
    if not text or is_mostly_korean(text):
        return text
    if not TRANSLATOR_AVAILABLE:
        return text
    try:
        chunk = text[:4500]
        translated = GoogleTranslator(source="auto", target="ko").translate(chunk)
        return translated if translated else text
    except Exception:
        return text


def make_summary(description: str, max_chars: int = SUMMARY_MAX_CHARS, translate: bool = True) -> str:
    if not description:
        return "(설명 없음)"
    text = re.sub(r"\s+", " ", description.strip())
    if len(text) > max_chars:
        text = text[:max_chars].rstrip() + "..."
    if translate:
        text = translate_to_korean(text)
    return text


# ---------------------------------------------------------------------------
# 키워드 필터
# ---------------------------------------------------------------------------

def title_passes_keyword_filter(title: str, include_keywords=None, exclude_keywords=None,
                                 include_mode: str = "OR") -> bool:
    title_lower = (title or "").lower()

    exclude_keywords = [k.strip().lower() for k in (exclude_keywords or []) if k.strip()]
    for kw in exclude_keywords:
        if kw in title_lower:
            return False

    include_keywords = [k.strip().lower() for k in (include_keywords or []) if k.strip()]
    if not include_keywords:
        return True

    if include_mode == "AND":
        return all(kw in title_lower for kw in include_keywords)
    return any(kw in title_lower for kw in include_keywords)


# ---------------------------------------------------------------------------
# 메인 수집 함수
# ---------------------------------------------------------------------------

def collect_channel_videos(channel_raw: str, api_key: str, start_date: datetime, end_date: datetime,
                            translate: bool = True, verbose: bool = True,
                            progress_callback=None,
                            include_keywords=None, exclude_keywords=None,
                            include_mode: str = "OR",
                            sort_by: str = "date",
                            check_shorts: bool = True):
    """
    채널 하나의 지정 기간 영상을 전부 수집해 리스트로 반환한다.
    각 영상 dict에 "type"(롱폼/숏폼/라이브) 필드가 채워진다.
    """
    def _log(msg):
        if verbose:
            print(msg)
        if progress_callback:
            progress_callback(msg)

    _log(f"[1/3] 채널 확인 중... ({channel_raw})")
    channel_info = resolve_channel_id(channel_raw, api_key)
    _log(f"      -> 채널 확인됨: {channel_info['title']}")

    _log("[2/3] 업로드 목록 조회 중...")
    video_ids = fetch_playlist_video_ids(
        channel_info["uploads_playlist_id"], api_key, start_date, end_date,
        progress_callback=progress_callback,
    )
    _log(f"      -> 총 {len(video_ids)}개 영상(기간 내) 발견.")

    if not video_ids:
        return []

    _log("[3/3] 영상 상세 정보 조회 중...")
    detail_items = fetch_videos_detail(video_ids, api_key)

    results = []
    skipped_by_keyword = 0

    for item in detail_items:
        snippet = item.get("snippet", {})
        content_details = item.get("contentDetails", {})
        statistics = item.get("statistics", {})
        live_details = item.get("liveStreamingDetails")

        title = snippet.get("title") or "(제목 없음)"

        if not title_passes_keyword_filter(title, include_keywords, exclude_keywords, include_mode):
            skipped_by_keyword += 1
            continue

        published_at_str = snippet.get("publishedAt")
        upload_date = datetime.fromisoformat(published_at_str.replace("Z", "+00:00")).replace(tzinfo=None)

        # 타입 판별: 라이브 우선, 그다음 실제 /shorts/ URL 확인
        if live_details is not None:
            video_type = "라이브"
        elif check_shorts:
            shorts_result = is_shorts_video(item["id"])
            if shorts_result is True:
                video_type = "숏폼"
            elif shorts_result is False:
                video_type = "롱폼"
            else:
                video_type = "확인불가"
        else:
            video_type = "확인불가"

        description = snippet.get("description") or ""

        results.append({
            "title": title,
            "upload_date": upload_date,
            "url": f"https://www.youtube.com/watch?v={item['id']}",
            "summary": make_summary(description, translate=translate),
            "type": video_type,
            "view_count": int(statistics.get("viewCount", 0)) if statistics.get("viewCount") else None,
            "like_count": int(statistics.get("likeCount", 0)) if statistics.get("likeCount") else None,
            "comment_count": int(statistics.get("commentCount", 0)) if statistics.get("commentCount") else None,
        })

        _log(f"      [{len(results)}] {upload_date.strftime('%Y-%m-%d')} | {video_type} | {title[:40]}")

    if skipped_by_keyword:
        _log(f"      -> 키워드 조건에 안 맞아 제외된 영상: {skipped_by_keyword}개")

    sort_key_map = {
        "views": lambda x: x.get("view_count") or 0,
        "likes": lambda x: x.get("like_count") or 0,
        "comments": lambda x: x.get("comment_count") or 0,
    }
    if sort_by in sort_key_map:
        results.sort(key=sort_key_map[sort_by], reverse=True)
    else:
        results.sort(key=lambda x: x["upload_date"], reverse=True)

    return results


def collect_multi_channel_videos(channel_raw_list, api_key: str, start_date: datetime, end_date: datetime,
                                  translate: bool = True, verbose: bool = True,
                                  progress_callback=None,
                                  include_keywords=None, exclude_keywords=None,
                                  include_mode: str = "OR",
                                  sort_by: str = "date",
                                  check_shorts: bool = True):
    """
    여러 채널을 순서대로 조회한다.
    """
    def _log(msg):
        if verbose:
            print(msg)
        if progress_callback:
            progress_callback(msg)

    channel_results = []

    for i, raw in enumerate(channel_raw_list, start=1):
        raw = raw.strip()
        if not raw:
            continue

        _log(f"\n=== 채널 {i}/{len(channel_raw_list)}: {raw} ===")

        try:
            videos = collect_channel_videos(
                raw, api_key, start_date, end_date,
                translate=translate, verbose=verbose, progress_callback=progress_callback,
                include_keywords=include_keywords, exclude_keywords=exclude_keywords,
                include_mode=include_mode, sort_by=sort_by, check_shorts=check_shorts,
            )
        except YouTubeAPIError as e:
            _log(f"      [오류] '{raw}' 채널 조회 실패: {e}")
            videos = []
        except Exception as e:
            _log(f"      [오류] '{raw}' 채널 조회 중 예상치 못한 문제: {e}")
            videos = []

        channel_results.append({
            "channel_input": raw,
            "videos": videos,
        })

    return channel_results


# ---------------------------------------------------------------------------
# 엑셀 생성
# ---------------------------------------------------------------------------

def _fmt_count(n):
    if n is None:
        return ""
    try:
        return f"{int(n):,}"
    except (TypeError, ValueError):
        return ""


def autosize_columns(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header(ws, ncols):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def write_sheet(ws, videos, show_channel_column: bool = False):
    if show_channel_column:
        headers = ["채널", "제목", "업로드 날짜", "구분", "조회수", "좋아요", "댓글수", "링크", "내용 요약"]
    else:
        headers = ["제목", "업로드 날짜", "구분", "조회수", "좋아요", "댓글수", "링크", "내용 요약"]
    ws.append(headers)

    body_font = Font(name="Arial", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="top")
    number_align = Alignment(horizontal="right", vertical="top")

    short_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    long_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    live_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    type_col_idx = 3 if show_channel_column else 2
    number_col_idxs = [4, 5, 6] if show_channel_column else [3, 4, 5]

    for v in videos:
        row = []
        if show_channel_column:
            row.append(v.get("channel_label", ""))
        row.extend([
            v["title"],
            v["upload_date"].strftime("%Y-%m-%d"),
            v["type"],
            _fmt_count(v.get("view_count")),
            _fmt_count(v.get("like_count")),
            _fmt_count(v.get("comment_count")),
            v["url"],
            v["summary"],
        ])
        ws.append(row)

    ncols = len(headers)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ncols):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap_align

        type_cell = row[type_col_idx]
        type_cell.alignment = center_align
        if type_cell.value == "숏폼":
            type_cell.fill = short_fill
        elif type_cell.value == "롱폼":
            type_cell.fill = long_fill
        elif type_cell.value == "라이브":
            type_cell.fill = live_fill

        for idx in number_col_idxs:
            row[idx].alignment = number_align

    style_header(ws, ncols=ncols)
    widths = [18, 40, 14, 10, 11, 10, 10, 40, 55] if show_channel_column else [40, 14, 10, 11, 10, 10, 40, 55]
    autosize_columns(ws, widths=widths)


def build_excel(videos, output):
    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "전체"
    write_sheet(ws_all, videos)

    for sheet_name, vtype in [("롱폼", "롱폼"), ("숏폼", "숏폼"), ("라이브", "라이브")]:
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, [v for v in videos if v["type"] == vtype])

    wb.save(output)


def _safe_sheet_name(name: str, used_names: set) -> str:
    cleaned = re.sub(r'[\\/*?:\[\]]', "_", name).strip() or "채널"
    cleaned = cleaned[:22]
    candidate = cleaned
    n = 2
    while candidate in used_names:
        candidate = f"{cleaned}_{n}"
        n += 1
    used_names.add(candidate)
    return candidate


def build_multi_channel_excel(channel_results, output):
    wb = Workbook()
    wb.remove(wb.active)

    used_names = set()
    all_videos_combined = []

    for cr in channel_results:
        label = cr["channel_input"]
        videos = cr["videos"]

        for v in videos:
            v["channel_label"] = label
        all_videos_combined.extend(videos)

        base_name = _safe_sheet_name(label, used_names)

        ws_all = wb.create_sheet(f"{base_name}_전체")
        write_sheet(ws_all, videos)

        for suffix, vtype in [("롱폼", "롱폼"), ("숏폼", "숏폼"), ("라이브", "라이브")]:
            ws = wb.create_sheet(f"{base_name}_{suffix}")
            write_sheet(ws, [v for v in videos if v["type"] == vtype])

    ws_combined = wb.create_sheet("모든채널_통합")
    write_sheet(ws_combined, all_videos_combined, show_channel_column=True)
    wb.move_sheet("모든채널_통합", offset=-(len(wb.sheetnames) - 1))

    wb.save(output)
